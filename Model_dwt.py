import numpy as np
import pandas as pd
from keras.layers import Layer
from keras.models import Model
from keras.utils import plot_model
from sklearn.model_selection import train_test_split
from keras.layers import Dense, Flatten, Conv1D, MaxPooling1D, Input, add
from keras import backend as K
from keras.layers import Multiply
import os
import shutil
from openpyxl import Workbook
import tensorflow as tf
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt
from keras.models import model_from_json
import Data_process_dwt

def load_total_data(path, num):
    df = pd.read_pickle(path)
    x = np.expand_dims(df.values[:, 0:-1].astype(float), axis=2)  # Adding a one-dimensional axis
    y = df.values[:, -1] / num
    # Divide training set, test set
    x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.2, shuffle=True)
    print("Loading of data complete!")
    return x_train, x_test, y_train, y_test


# Loading data
def load_distribute_data(all_path, num):
    x_train_ = []
    x_test_ = []
    y_train_ = []
    y_test_ = []
    # Loading data
    for i in all_path:
        df = pd.read_pickle(i)
        x = np.expand_dims(df.values[:, 0:-1].astype(float), axis=2)  # Adding a one-dimensional axis
        y = df.values[:, -1] / num
        # Divide training set, test set
        x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.05, shuffle=True)
        x_train_.append(x_train)
        x_test_.append(x_test)
        y_train_.append(y_train)
        y_test_.append(y_test)
    print("Loading of data complete!")
    return x_train_, x_test_, y_train_[0], y_test_[0]


# Custom metric function, determination factor R_Squares
def coeff_determination(y_true, y_pred):
    SS_res = K.sum(K.square(y_true - y_pred))
    SS_tot = K.sum(K.square(y_true - K.mean(y_true)))
    return (1 - SS_res / (SS_tot + K.epsilon()))


class MyLayer(Layer):

    def __init__(self, output_dim, **kwargs):
        self.output_dim = output_dim
        super(MyLayer, self).__init__(**kwargs)

    def build(self, input_layer):
        # Create a trainable weight for this layer
        self.kernel = self.add_weight(name='kernel',
                                      shape=(1, self.output_dim),
                                      initializer='uniform',
                                      trainable=True)
        super(MyLayer, self).build(input_layer)

    def call(self, input_layer):
        return tf.tile(self.kernel, [tf.shape(input_layer)[0], 1])

    def get_config(self):
        config = super().get_config().copy()
        config.update({
            'output_dim': self.output_dim,
        })
        return config


# # Defining the structure of a neural network
# def build_CNN_model(model_structure):
#     input1 = Input(shape=(570, 1))
#     conv_layer1_1 = Conv1D(16, 3, strides=1, activation='relu')(input1)
#     conv_layer1_1_ = Conv1D(16, 3, strides=1, activation='relu')(conv_layer1_1)
#     max_layer1_1 = MaxPooling1D(3)(conv_layer1_1_)
#     conv_layer1_2 = Conv1D(32, 3, strides=1, activation='relu')(max_layer1_1)
#     conv_layer1_2_ = Conv1D(32, 3, strides=1, activation='relu')(conv_layer1_2)
#     max_layer1_2 = MaxPooling1D(3)(conv_layer1_2_)
#     conv_layer1_3 = Conv1D(32, 3, activation='relu')(max_layer1_2)
#     conv_layer1_3_ = Conv1D(32, 3, activation='relu')(conv_layer1_3)
#     max_layer1_3 = MaxPooling1D(3)(conv_layer1_3_)
#     conv_layer1_4 = Conv1D(32, 3, activation='relu')(max_layer1_3)
#     conv_layer1_4_ = Conv1D(32, 3, activation='relu')(conv_layer1_4)
#     max_layer1_4 = MaxPooling1D(3)(conv_layer1_4_)
#     flatten = Flatten()(max_layer1_4)
#     f1 = Dense(1, activation='linear', name='prediction_one')(flatten)
#     model = Model(outputs=f1, inputs=input1)
#     model.summary()
#     plot_model(model, to_file=model_structure, show_shapes=True)  # Printed model structure
#     return model

# Defining the structure of a neural network
def build_CNN_model(model_structure):
    input1 = Input(shape=(570, 1))
    conv_layer1_1 = Conv1D(16, 3, strides=1, activation='relu')(input1)
    # conv_layer1_1_ = Conv1D(16, 3, strides=1, activation='relu')(conv_layer1_1)
    max_layer1_1 = MaxPooling1D(3)(conv_layer1_1)
    conv_layer1_2 = Conv1D(32, 3, strides=1, activation='relu')(max_layer1_1)
    # conv_layer1_2_ = Conv1D(32, 3, strides=1, activation='relu')(conv_layer1_2)
    max_layer1_2 = MaxPooling1D(3)(conv_layer1_2)
    conv_layer1_3 = Conv1D(32, 3, activation='relu')(max_layer1_2)
    # conv_layer1_3_ = Conv1D(32, 3, activation='relu')(conv_layer1_3)
    max_layer1_3 = MaxPooling1D(3)(conv_layer1_3)
    conv_layer1_4 = Conv1D(32, 3, activation='relu')(max_layer1_3)
    # conv_layer1_4_ = Conv1D(32, 3, activation='relu')(conv_layer1_4)
    max_layer1_4 = MaxPooling1D(3)(conv_layer1_4)
    flatten = Flatten()(max_layer1_4)
    f1 = Dense(1, activation='linear', name='prediction_one')(flatten)
    model = Model(outputs=f1, inputs=input1)
    model.summary()
    plot_model(model, to_file=model_structure, show_shapes=True)  # Printed model structure
    return model



def build_DPSRNN_model(model_structure):
    # Input can be integrated at any layer, here we only integrate the input layer
    input1 = Input(shape=(456, 1))
    input2 = Input(shape=(37, 1))
    input3 = Input(shape=(38, 1))
    input4 = Input(shape=(39, 1))
    conv_layer1_1 = Conv1D(16, 3, strides=2, activation='relu')(input1)
    max_layer1_1 = MaxPooling1D(2)(conv_layer1_1)
    conv_layer1_2 = Conv1D(32, 4, strides=2, activation='relu')(max_layer1_1)
    max_layer1_2 = MaxPooling1D(2)(conv_layer1_2)
    conv_layer1_3 = Conv1D(32, 4, activation='relu')(max_layer1_2)
    max_layer1_3 = MaxPooling1D(3)(conv_layer1_3)
    w_batch1 = MyLayer(output_dim=32)(max_layer1_3)
    x1 = Multiply()([w_batch1, max_layer1_3])

    conv_layer2_1 = Conv1D(16, 3, activation='relu')(input2)
    max_layer2_1 = MaxPooling1D(2)(conv_layer2_1)
    conv_layer2_2 = Conv1D(32, 2, activation='relu')(max_layer2_1)
    max_layer2_2 = MaxPooling1D(2)(conv_layer2_2)
    w_batch2 = MyLayer(output_dim=32)(max_layer2_2)
    x2 = Multiply()([w_batch2, max_layer2_2])

    conv_layer3_1 = Conv1D(16, 3, activation='relu')(input3)
    max_layer3_1 = MaxPooling1D(2)(conv_layer3_1)
    conv_layer3_2 = Conv1D(32, 3, activation='relu')(max_layer3_1)
    max_layer3_2 = MaxPooling1D(2)(conv_layer3_2)
    w_batch3 = MyLayer(output_dim=32)(max_layer3_2)
    x3 = Multiply()([w_batch3, max_layer3_2])

    conv_layer4_1 = Conv1D(16, 3, activation='relu')(input4)
    max_layer4_1 = MaxPooling1D(2)(conv_layer4_1)
    conv_layer4_2 = Conv1D(32, 3, activation='relu')(max_layer4_1)
    max_layer4_2 = MaxPooling1D(2)(conv_layer4_2)
    w_batch4 = MyLayer(output_dim=32)(max_layer4_2)
    x4 = Multiply()([w_batch4, max_layer4_2])

    Added = add([x1, x2, x3, x4])
    conv_layer5_1 = Conv1D(16, 3, padding='same', activation='relu')(Added)
    max_layer5_1 = MaxPooling1D(2)(conv_layer5_1)
    flatten = Flatten()(max_layer5_1)
    f1 = Dense(1, activation='linear', name='prediction_one')(flatten)
    model = Model(outputs=f1, inputs=[input1, input2, input3, input4])
    model.summary()
    plot_model(model, to_file=model_structure, show_shapes=True)  # Printed model structure
    return model

def test_build_DPSRNN_model():
    # Input can be integrated at any layer, here we only integrate the input layer
    input1 = Input(shape=(456, 1))
    input2 = Input(shape=(37, 1))
    input3 = Input(shape=(38, 1))
    input4 = Input(shape=(39, 1))
    conv_layer1_1 = Conv1D(16, 3, strides=2, activation='relu')(input1)
    max_layer1_1 = MaxPooling1D(2)(conv_layer1_1)
    conv_layer1_2 = Conv1D(32, 4, strides=2, activation='relu')(max_layer1_1)
    max_layer1_2 = MaxPooling1D(2)(conv_layer1_2)
    conv_layer1_3 = Conv1D(32, 4, activation='relu')(max_layer1_2)
    max_layer1_3 = MaxPooling1D(3)(conv_layer1_3)
    w_batch1 = MyLayer(output_dim=32)(max_layer1_3)
    x1 = Multiply()([w_batch1, max_layer1_3])

    conv_layer2_1 = Conv1D(16, 3, activation='relu')(input2)
    max_layer2_1 = MaxPooling1D(2)(conv_layer2_1)
    conv_layer2_2 = Conv1D(32, 2, activation='relu')(max_layer2_1)
    max_layer2_2 = MaxPooling1D(2)(conv_layer2_2)
    w_batch2 = MyLayer(output_dim=32)(max_layer2_2)
    x2 = Multiply()([w_batch2, max_layer2_2])

    conv_layer3_1 = Conv1D(16, 3, activation='relu')(input3)
    max_layer3_1 = MaxPooling1D(2)(conv_layer3_1)
    conv_layer3_2 = Conv1D(32, 3, activation='relu')(max_layer3_1)
    max_layer3_2 = MaxPooling1D(2)(conv_layer3_2)
    w_batch3 = MyLayer(output_dim=32)(max_layer3_2)
    x3 = Multiply()([w_batch3, max_layer3_2])

    conv_layer4_1 = Conv1D(16, 3, activation='relu')(input4)
    max_layer4_1 = MaxPooling1D(2)(conv_layer4_1)
    conv_layer4_2 = Conv1D(32, 3, activation='relu')(max_layer4_1)
    max_layer4_2 = MaxPooling1D(2)(conv_layer4_2)
    w_batch4 = MyLayer(output_dim=32)(max_layer4_2)
    x4 = Multiply()([w_batch4, max_layer4_2])

    Added = add([x1, x2, x3, x4])
    conv_layer5_1 = Conv1D(16, 3, padding='same', activation='relu')(Added)
    max_layer5_1 = MaxPooling1D(2)(conv_layer5_1)
    flatten = Flatten()(max_layer5_1)
    f1 = Dense(1, activation='linear', name='prediction_one')(flatten)
    model = Model(outputs=f1, inputs=[input1, input2, input3, input4])
    model.summary()
    # plot_model(model, to_file=model_structure, show_shapes=True)  # Printed model structure
    return model

def test_build_CNN_model():
    input1 = Input(shape=(570, 1))
    conv_layer1_1 = Conv1D(16, 3, strides=1, activation='relu')(input1)
    # conv_layer1_1_ = Conv1D(16, 3, strides=1, activation='relu')(conv_layer1_1)
    max_layer1_1 = MaxPooling1D(3)(conv_layer1_1)
    conv_layer1_2 = Conv1D(32, 3, strides=1, activation='relu')(max_layer1_1)
    # conv_layer1_2_ = Conv1D(32, 3, strides=1, activation='relu')(conv_layer1_2)
    max_layer1_2 = MaxPooling1D(3)(conv_layer1_2)
    conv_layer1_3 = Conv1D(32, 3, activation='relu')(max_layer1_2)
    # conv_layer1_3_ = Conv1D(32, 3, activation='relu')(conv_layer1_3)
    max_layer1_3 = MaxPooling1D(3)(conv_layer1_3)
    conv_layer1_4 = Conv1D(32, 3, activation='relu')(max_layer1_3)
    # conv_layer1_4_ = Conv1D(32, 3, activation='relu')(conv_layer1_4)
    max_layer1_4 = MaxPooling1D(3)(conv_layer1_4)
    flatten = Flatten()(max_layer1_4)
    f1 = Dense(1, activation='linear', name='prediction_one')(flatten)
    model = Model(outputs=f1, inputs=input1)
    model.summary()
    # plot_model(model, to_file=model_structure, show_shapes=True)  # Printed model structure
    return model

# Save prediction results
def save_DPSRNN_excel(predicted, Y_test, name, num):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    Y_test = Y_test
    ws.cell(1, 1, 'pre_result')
    ws.cell(1, 2, 'ori_result')
    for i in range(len(predicted)):
        ws.cell(i + 2, 1, predicted[i][0] * num)
        ws.cell(i + 2, 2, Y_test[i] * num)
    wb.save(name)
    print("Save value to finish!")


def save_DPSRNN_loss(history, name):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    hist = pd.DataFrame(history.history)
    loss = hist['loss']
    val_loss = hist['val_loss']
    ws.cell(1, 1, 'loss')
    ws.cell(1, 2, 'val_loss')
    for i in range(len(loss)):
        ws.cell(i + 2, 1, loss[i])
        ws.cell(i + 2, 2, val_loss[i])
    wb.save(name)
    print("Save loss to finish!")


# Save prediction results
def save_cnn_excel(predicted, y_test, name, num):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    ws.cell(1, 1, 'pre_result')
    ws.cell(1, 2, 'ori_result')
    for i in range(len(predicted)):
        ws.cell(i + 2, 1, predicted[i][0] * num)
        ws.cell(i + 2, 2, y_test[i] * num)
    wb.save(name)
    print("Save value to finish!")


def save_cnn_loss(history, name):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    hist = pd.DataFrame(history.history)
    loss = hist['loss']
    val_loss = hist['val_loss']
    ws.cell(1, 1, 'loss')
    ws.cell(1, 2, 'val_loss')
    for i in range(len(loss)):
        ws.cell(i + 2, 1, loss[i])
        ws.cell(i + 2, 2, val_loss[i])
    wb.save(name)
    print("Save loss to finish!")


def save_mae_r2(mae, r2, name):
    wb = Workbook()  # Create a new Excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active

    ws.cell(1, 1, 'MAE')
    ws.cell(1, 2, 'R2')
    ws.cell(2, 1, mae)
    ws.cell(2, 2, r2)
    wb.save(name)
    print("Save mae & r2 to finish!")


# Predicted data
def predict_data(model, X_test, optimizer, loss):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = model.predict(X_test)
    return predicted


# Calculation of the decision factor
def show_scores(predicted, Y_test):
    r2_scores = r2_score(predicted, Y_test)
    print("R2:", r2_scores)
    return r2_scores



# Calculating the mean absolute error
def mean_error(predicted, y_test, num):
    y_test_size = np.reshape(y_test, (len(y_test), 1))
    result = np.mean(abs(predicted * num - y_test_size * num))
    print("MAE:", result)
    return result


# Preservation of models
def save_model(model, name1, name2):
    # Convert their model grid structure to json storage
    # Store model parameter weights as h5 files
    model_json = model.to_json()
    with open(name1, 'w') as json_file:
        json_file.write(model_json)
    model.save_weights(name2)
    print("Save model complete!")


def train_CNN_model(model, name, loss_name, model_para, model_stru, optimizer, loss, X_train, X_test, Y_train, Y_test,
                    name1, name2, mae_r2_name):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    history = model.fit(X_train, Y_train,
                        batch_size=128,
                        epochs=100,
                        validation_data=(X_test, Y_test)
                        )
    predicted = predict_data(model, X_test, optimizer=optimizer, loss=loss)
    save_model(model, model_stru, model_para)
    mae = mean_error(predicted, Y_test, num)
    r2 = show_scores(predicted, Y_test)
    save_cnn_excel(predicted, Y_test, name, num)
    plot_history(history, name1, name2)
    save_cnn_loss(history, loss_name)
    save_mae_r2(mae, r2, mae_r2_name)


def train_DPSRNN_model(model, name, loss_name, model_para, model_stru, optimizer, loss, X_train, X_test, Y_train, Y_test,
                     name1, name2, mae_r2_name):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    history = model.fit(X_train, Y_train,
                        batch_size=128,
                        epochs=1000,
                        validation_data=(X_test, Y_test)
                        )
    predicted = predict_data(model, X_test, optimizer=optimizer, loss=loss)
    save_model(model, model_stru, model_para)
    mae = mean_error(predicted, Y_test, num)
    r2 = show_scores(predicted, Y_test)
    save_DPSRNN_excel(predicted, Y_test, name, num)
    plot_history(history, name1, name2)
    save_DPSRNN_loss(history, loss_name)
    save_mae_r2(mae, r2, mae_r2_name)



def del_files(path):
    if os.path.exists(path):
        shutil.rmtree(path, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")


def plot_history(history, name1, name2):
    hist = pd.DataFrame(history.history)
    hist['epoch'] = history.epoch
    plt.figure()
    plt.xlabel('Epoch')
    plt.ylabel('loss')
    plt.plot(hist['epoch'], hist['loss'],
             label='Train loss')
    plt.plot(hist['epoch'], hist['val_loss'],
             label='Val loss')
    plt.ylim([-0.001, 0.05])
    plt.legend()
    plt.savefig(name1, dpi=600)
    plt.figure()
    plt.xlabel('Epoch')
    plt.ylabel('coeff_determination')
    plt.plot(hist['epoch'], hist['coeff_determination'],
             label='Train coeff_determination')
    plt.plot(hist['epoch'], hist['val_coeff_determination'],
             label='Val coeff_determination')
    plt.ylim([-0.5, 1.5])
    plt.legend()
    plt.savefig(name2, dpi=600)
    # plt.show()

# Pre-loading
def load_first_DPSRNN(model_stru2, model_para2):
    loaded_model = test_build_DPSRNN_model()
    loaded_model.load_weights(model_para2)
    print("Loading of pre-trained models complete!")
    return loaded_model

def load_first_cnn(model_stru2, model_para1):
    loaded_model = test_build_CNN_model()
    loaded_model.load_weights(model_para1)
    print("Loading of pre-trained models complete!")
    return loaded_model

# Save the test file
def save_test_excel(predicted, Y_test, name, num):
    wb = Workbook()  # Create a new excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    Y_test = Y_test
    ws.cell(1, 1, 'pre_result')
    ws.cell(1, 2, 'ori_result')
    for i in range(len(predicted)):
        ws.cell(i + 2, 1, predicted[i][0] * num)
        ws.cell(i + 2, 2, Y_test[i] * num)
    wb.save(name)
    print("Save excel to finish!")


def test_model_DPSRNN(path_all, model_stru2, model_para2, test_path, optimizer, loss, num):
    x_test_ = []
    y_test_ = []
    # Loading data
    for i in path_all:
        df = pd.read_pickle(i)
        X_test = np.expand_dims(df.values[:, 0:-1].astype(float), axis=2)  # Adding a one-dimensional axis
        Y_test = df.values[:, -1] / num
        x_test_.append(X_test)
        y_test_.append(Y_test)
    print("Loading of data complete!")
    x_test = x_test_
    y_test = y_test_[0]
    loaded_model = load_first_DPSRNN(model_stru2, model_para2)
    loaded_model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = predict_data(loaded_model, x_test, optimizer, loss)
    mean_error(predicted, y_test, num)
    # show_scores(predicted, y_test)
    print("Test complete!")
    save_test_excel(predicted, y_test, test_path, num)

def test_model_cnn(path, model_stru1, model_para1, test_path, optimizer, loss, num):

    # Loading data
    df = pd.read_pickle(path)
    x_test = np.expand_dims(df.values[:, 0:-1].astype(float), axis=2)  # Adding a one-dimensional axis
    y_test = df.values[:, -1] / num
    # Divide training set, test set
    print("Loading of data complete!")
    x_test = x_test
    y_test = y_test
    loaded_model = load_first_cnn(model_stru1, model_para1)
    loaded_model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = predict_data(loaded_model, x_test, optimizer, loss)

    mean_error(predicted, y_test, num)
    # show_scores(predicted, y_test)
    print("Test complete!")
    save_test_excel(predicted, y_test, test_path, num)

if __name__ == '__main__':

    path_x = Data_process_dwt.run()
    path_t = "Test_dwt"

    num = 1000
    optimizer = "adam"
    loss = "mean_squared_error"

    path0 = path_x + "/line_peak_dwt.pkl"
    path1 = path_x + "/lines_dwt.pkl"
    path2 = path_x + "/peak_1_dwt.pkl"
    path3 = path_x + "/peak_2_dwt.pkl"
    path4 = path_x + "/peak_3_dwt.pkl"
    path5 = path_x + "/peaks_dwt.pkl"

    dir1 = "Results_CNN_dwt"
    dir2= "Results_DPSRNN_dwt"

    test_path_cnn = path_t +"/test_cnn_dwt.xlsx"
    test_path_DPSRNN= path_t +"/test_DPSRNN_dwt.xlsx"

    model_structure1 = dir1 + "/CNN.png"
    model_para1 = dir1 + "/CNN.h5"
    model_stru1 = dir1 + "/CNN.json"
    model_pre_results1 = dir1 + "/CNN.xlsx"
    loss_name1 = dir1 + "/loss.xlsx"
    loss_png_name1 = dir1 + "/loss.png"
    R2_png_name1 = dir1 + "/R2.png"
    mae_r2_name1 = dir1 + "/mae_r2.xlsx"

    model_structure2 = dir2 + "/DPSRNN.png"
    model_para2 = dir2 + "/DPSRNN.h5"
    model_stru2 = dir2 + "/DPSRNN.json"
    model_pre_results2 = dir2 + "/DPSRNN.xlsx"
    loss_name2 = dir2 + "/loss.xlsx"
    loss_png_name2 = dir2 + "/loss.png"
    R2_png_name2 = dir2 + "/R2.png"
    mae_r2_name2 = dir2 + "/mae_r2.xlsx"


    # test_path1 = path_t +"/test_lines_dwt.pkl"
    # test_path2 = path_t +"/test_peak1_dwt.pkl"
    # test_path3 = path_t +"/test_peak2_dwt.pkl"
    # test_path4 = path_t +"/test_peak3_dwt.pkl"
    # test_path5 = path_t +"/test_line_peak_dwt.pkl"


    # measurement accuracy
    test_path1 = path_x + "/lines_dwt.pkl"
    test_path2 = path_x + "/peak_1_dwt.pkl"
    test_path3 = path_x + "/peak_2_dwt.pkl"
    test_path4 = path_x + "/peak_3_dwt.pkl"
    test_path5 = path_x + "/line_peak_dwt.pkl"


    dir_t = "Test_human_deal_dwt"
    test_path_10 = dir_t + "/test_DPSRNN_dwt.xlsx"


    Train_DPSRNN_model = True
    istest = False
    if istest:
        del_files(dir_t)
        os.mkdir(dir_t)
        data_path_DPSRNN = [test_path1, test_path2, test_path3, test_path4]
        # test_model_DPSRNN(data_path_DPSRNN, model_stru2, model_para2, test_path_10, optimizer, loss, num)
        # test_model_cnn(test_path5, model_stru1, model_para1, test_path_cnn, optimizer, loss, num)
    else:
        if Train_DPSRNN_model:
            del_files(dir2)
            os.mkdir(dir2)
            data_path = [path1, path2, path3, path4]
            X_train, X_test, Y_train, Y_test = load_distribute_data(data_path, num)
            model = build_DPSRNN_model(model_structure2)
            train_DPSRNN_model(model, model_pre_results2, loss_name2, model_para2, model_stru2, optimizer, loss, X_train,
                             X_test, Y_train, Y_test, loss_png_name2, R2_png_name2, mae_r2_name2)
        else:
            del_files(dir1)
            os.mkdir(dir1)
            X_train, X_test, Y_train, Y_test = load_total_data(path0, num)
            model = build_CNN_model(model_structure1)
            train_DPSRNN_model(model, model_pre_results1, loss_name1, model_para1, model_stru1, optimizer, loss, X_train,
                            X_test, Y_train, Y_test, loss_png_name1, R2_png_name1, mae_r2_name1)
