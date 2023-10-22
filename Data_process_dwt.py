import os
import random
import numpy as np
import xlwt
import re
import pandas as pd
from openpyxl import Workbook
import shutil
from decimal import Decimal
from pandas import DataFrame
import pywt


def read_data(path, newpath, xlsx_data, txt_num):
    '''
    * read_Data.py file is placed under the same root directory as the read data folder.
    * path：Enter the read data folder path.
    * Read the data folder layout as shown in the example.
    * After rerunning to read the data, if you rerun to read, you need to delete the newly generated **_ok folder in the Data folder before starting the operation.
    '''
    os.mkdir(newpath)
    os.mkdir(xlsx_data)
    path = path
    file_name_lists = []
    for file_name in os.listdir(path):
        file_name_lists.append(file_name)
    for file_name_list in file_name_lists:
        root_ = []
        dirs_ = []
        a = 0
        for root, dirs, files in os.walk(path + '\\' + file_name_list):
            root_.append(root)
            dirs_.append(dirs)
            dirs__ = dirs_[0]
        file_name_after = newpath + '\\' + file_name_list + '_ok'
        if not os.path.exists(file_name_after):
            os.mkdir(file_name_after)
        for i in root_[1:]:
            file_name_after_ = file_name_after + '\\' + dirs__[a]
            if not os.path.exists(file_name_after_):
                os.mkdir(file_name_after_)
            txt_ = []
            for file_name_ in os.listdir(i):
                txt_.append(i + '\\' + file_name_)
            txt_num = txt_num  # 取多少文件
            txt_ = txt_[-txt_num:]
            num_all = []
            for txt_name in txt_:
                contents = []
                with open(txt_name, 'r') as f:
                    for line in f.readlines():
                        line = line.split('\n')
                        line = line[0].split('\t')
                        line = list(map(float, line))
                        contents.append(line)
                for content in contents:
                    num_all.append("%.4f" % (float(content[1]) / txt_num))
                if len(num_all) > len(contents):
                    for ii in range(len(num_all)):
                        if ii < len(contents):
                            num_all[ii] = "%.4f" % (num_all[ii] + float(num_all[ii + len(contents)]))
                        else:
                            num_all.pop()
                num_all = list(map(float, num_all))
                f.close()
                txt_name_after = newpath + '\\' + file_name_list + '_ok\\' + dirs__[a] + "\\" + dirs__[a] + ".txt"
                with open(txt_name_after, "w") as ff:
                    for li in num_all:
                        ff.write(str(li) + "\n")
                ff.close()
            a += 1
        print(file_name_list, "Data reading completed！")
    print("All data read completed！")


def remove_bd(newpath):
    path = newpath
    nn_ = []
    file_name_lists = []
    for file_name in os.listdir(path):
        file_name_lists.append(file_name)
    for file_name_list in file_name_lists:
        n_p = path + '\\' + file_name_list + '_z'
        nn_.append(n_p)
        os.mkdir(n_p)
        root_ = []
        dirs_ = []

        for root, dirs, files in os.walk(path + '\\' + file_name_list):
            root_.append(root)
            dirs_.append(dirs)
            dirs__ = dirs_[0]
        root_.pop(0)
        root__ = root_[-1]
        dirs___ = dirs__[-1]
        root_.pop()
        dirs__.pop()
        bd_name = root__ + '\\' + dirs___ + '.txt'
        for i in range(len(root_)):
            data = []
            file_name = root_[i] + '\\' + dirs__[i] + '.txt'
            file_name_ = n_p + '\\' + dirs__[i] + '.txt'
            with open(bd_name) as bd_f:
                bd_file = bd_f.read().split('\n')
            bd_f.close()
            with open(file_name, 'r+') as f:
                file = f.read().split('\n')
            f.close()
            del (bd_file[-1])
            del (file[-1])
            bd_file = list(map(float, bd_file))
            file = list(map(float, file))
            for i in range(len(bd_file)):
                i_num = "%.4f" % ((file[i]) / (bd_file[i]))
                data.append(i_num)
            with open(file_name_, 'w') as f_:
                f_.truncate(0)
                for ii in data:
                    f_.write(ii + "\n")
            f_.close()
    print('Successfully removed the backing!')
    return nn_


def writeinexcel(path, nn):
    path = path
    le_ = 0
    wb1 = xlwt.Workbook(encoding='utf-8')  # Create a new excel file
    w1 = wb1.add_sheet('one')  # Add a new table with the name first
    ipath_ = path[0]
    file_name_lists_ = []
    file_name_lists = []
    for file_name_ in os.listdir(ipath_):
        file_name_ = re.sub('\D', '', file_name_)
        b = list(file_name_)
        ans = "".join(map(str, b))
        file_name_lists_.append(ans)
    file_name_lists_ = list(map(int, file_name_lists_))
    file_name_lists_.sort()
    le = len(file_name_lists_)
    for le_i in range(le):
        w1.write(0, le_i + le_, file_name_lists_[le_i])
    for file_name_ in os.listdir(ipath_):
        file_name_lists.append(file_name_)
    file_name_lists.sort(key=lambda x: int(x[:-nn]))
    for i_a in range(len(file_name_lists)):
        path_ = ipath_ + '\\' + file_name_lists[i_a]
        ii = 1
        for line in open(path_, encoding='utf-8'):
            if line == '\n':
                continue
            else:
                w1.write(ii, i_a, float(line))
                ii += 1

    p1 = path[0] + ".xls"
    wb1.save(p1)

    return p1


def del_files(path, name1):
    if os.path.exists(path):
        shutil.rmtree(path, ignore_errors=False, onerror=None)
    if os.path.exists(name1):
        shutil.rmtree(name1, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")


def koumanbian(x, data_all, a, b, e, f, i, j, m, n):
    aa = []
    j_ = 0

    line1 = list(range(a - 1, b))
    # print("line1:",len(line1))
    line2 = list(range(e - 1, f))
    # print("line2:", len(line2))
    line3 = list(range(i - 1, j))
    # print("line3:", len(line3))
    line4 = list(range(m - 1, n))
    # print("line4:", len(line4))
    line_z = list(range(a - 1, n))
    # print(len(line_z))

    listall = line1 + line2 + line3 + line4
    # print(len(listall))

    for i in x:
        zz1 = np.polyfit(listall, i, 3)  # 拟合
        pp1 = np.poly1d(zz1)
        jj = data_all[j_]

        # aa.append(i/pp1([i for i in range(len(i))]))
        aa.append(np.log(jj / pp1(line_z)))
        # aa.append(jj / pp1(line_z))
        j_ += 1

    return np.array(aa)


def dif(path):
    # Weak characteristic absorption region 1
    a = 627
    b = 700

    # Characteristic absorption peak 1
    c = 701
    d = 737

    # Weak characteristic absorption region 2
    e = 738
    f = 888

    # Characteristic absorption peak 2
    g = 889
    h = 926

    # Weak characteristic absorption region 3
    i = 927
    j = 1106

    # Characteristic absorption peak 3
    k = 1107
    l = 1145

    # Weak characteristic absorption region 4
    m = 1146
    n = 1196

    data = pd.read_excel(path)
    columns = data.columns
    wb_all = Workbook()
    wb_all.create_sheet(index=0, title="all")
    ws_all = wb_all.active

    data_line_1 = data.iloc[a - 1:b, :]
    data_peak_1 = data.iloc[c - 1:d, :]
    data_line_2 = data.iloc[e - 1:f, :]
    data_peak_2 = data.iloc[g - 1:h, :]
    data_line_3 = data.iloc[i - 1:j, :]
    data_peak_3 = data.iloc[k - 1:l, :]
    data_line_4 = data.iloc[m - 1:n, :]

    data_line_1 = np.array(data_line_1)
    data_peak_1 = np.array(data_peak_1)
    data_line_2 = np.array(data_line_2)
    data_peak_2 = np.array(data_peak_2)
    data_line_3 = np.array(data_line_3)
    data_peak_3 = np.array(data_peak_3)
    data_line_4 = np.array(data_line_4)

    line_1_shape = data_line_1.shape[0]
    line_2_shape = data_line_2.shape[0]
    line_3_shape = data_line_3.shape[0]
    line_4_shape = data_line_4.shape[0]
    peak_1_shape = data_peak_1.shape[0]
    peak_2_shape = data_peak_2.shape[0]
    peak_3_shape = data_peak_3.shape[0]

    data_all = np.concatenate(
        (data_line_1, data_peak_1, data_line_2, data_peak_2, data_line_3, data_peak_3, data_line_4), axis=0)
    # print(data_all.shape)
    data_all = pd.DataFrame(data_all, columns=columns)
    data_all = data_all.T
    data_all = np.array(data_all)
    # print(data_all.shape)
    data_line_all = np.concatenate((data_line_1, data_line_2, data_line_3, data_line_4), axis=0)
    data_line_all = pd.DataFrame(data_line_all, columns=columns)
    data_line_all = data_line_all.T
    data_line_all = np.array(data_line_all)
    # print(data_line_all.shape)
    # print(data_all.shape)
    data_deal_all_line = koumanbian(data_line_all, data_all, a, b, e, f, i, j, m, n)
    deal_data_line_all = pd.DataFrame(data_deal_all_line)
    # print("deal_data_line_all",deal_data_line_all)
    columns_ = deal_data_line_all.columns
    # all_data
    l = len(columns_)
    i1 = 0
    for i in range(l):
        i1 += 1
        lie = deal_data_line_all[columns_[i]]
        for j in range(len(lie)):
            ws_all.cell(1, i + 1, i1)
            ws_all.cell(j + 2, i + 1, lie[j])
    for col_i in range(len(columns)):
        col = columns[col_i]
        ws_all.cell(col_i + 2, l + 1, col)
    ws_all.cell(1, l + 1, l + 1)
    name = path + '.xlsx'
    wb_all.save(name)
    print("Data differential completion!")
    return name, line_1_shape, line_2_shape, line_3_shape, line_4_shape, peak_1_shape, peak_2_shape, peak_3_shape


def Extended_data(path, end_file_name0):
    wb = Workbook()
    wb.create_sheet(index=0, title="all")
    ws = wb.active

    data1 = pd.read_excel(path)
    columns1 = data1.columns
    le = 0
    for i in range(data1.shape[1]):
        lie1 = data1[columns1[i]]
        data_1 = lie1
        ws.cell(1, i + 1, str(columns1[i]))
        for ii in range(len(data_1)):
            ws.cell(ii + 2, i + 1, data_1[ii])
    le += int(data1.shape[0])
    index1 = data1.index
    for time in range(40):
        a = random.uniform(0, 1.25)
        a = round(a, 3)
        b = random.uniform(0, 1.25)
        b = round(b, 3)
        if a == 0 or b == 0:
            a += 0.1
            b += 0.1
        for i in range(data1.shape[0]):
            for j in range(data1.shape[0]):
                ind1 = list(data1.loc[index1[i]])
                # ind1.pop(0)
                ind2 = list(data1.loc[index1[j]])
                # ind2.pop(0)
                ind1 = np.array(ind1)
                ind2 = np.array(ind2)
                '''Combination Rules'''
                a = float(a)
                b = float(b)
                data_1 = a * ind1 + b * ind2
                for iii in range(len(data_1)):
                    ws.cell(le + 2 + j, iii + 1, data_1[iii])
            le += int(data1.shape[0])
        print("circulate", time, "time")
    wb.save(end_file_name0)
    path = to_pkl(end_file_name0)
    return path


def to_pkl(path):
    # Read excel files
    df1 = DataFrame(pd.read_excel(path))
    dir_name = os.path.dirname(path)
    base_name = os.path.basename(path)
    suffix = base_name.split(".")[0]
    path_ = dir_name + "/" + suffix + ".pkl"
    df1.to_pickle(path_)
    return path_


def Distribute_data(path5, end_file_name2, end_file_name3, end_file_name3_1, end_file_name3_2,
                    end_file_name3_3, line_1_shape, line_2_shape, line_3_shape, line_4_shape, peak_1_shape,
                    peak_2_shape, peak_3_shape):
    wb1 = Workbook()
    wb1.create_sheet(index=0, title="all")
    ws1 = wb1.active

    wb2 = Workbook()
    wb2.create_sheet(index=0, title="all")
    ws2 = wb2.active

    wb2_1 = Workbook()
    wb2_1.create_sheet(index=0, title="all")
    ws2_1 = wb2_1.active

    wb2_2 = Workbook()
    wb2_2.create_sheet(index=0, title="all")
    ws2_2 = wb2_2.active

    wb2_3 = Workbook()
    wb2_3.create_sheet(index=0, title="all")
    ws2_3 = wb2_3.active

    path = path5
    data = pd.read_pickle(path)

    columns = data.columns
    cos = data[columns[-1]]

    # line_data
    data_deal_line = pd.DataFrame(data)
    data_deal_line_1 = data_deal_line.iloc[:, 0:line_1_shape]

    data_deal_line_2 = data_deal_line.iloc[:, line_1_shape + peak_1_shape:line_1_shape + line_2_shape + peak_1_shape]
    data_deal_line_3 = data_deal_line.iloc[:,
                       line_1_shape + line_2_shape + peak_1_shape + peak_2_shape:line_1_shape + line_2_shape + line_3_shape + +peak_1_shape + peak_2_shape]
    data_deal_line_4 = data_deal_line.iloc[:,
                       line_1_shape + line_2_shape + line_3_shape + peak_1_shape + peak_2_shape + peak_3_shape:line_1_shape + line_2_shape + line_3_shape + line_4_shape + peak_1_shape + peak_2_shape + peak_3_shape]
    data_deal_line_1 = np.array(data_deal_line_1)
    data_deal_line_2 = np.array(data_deal_line_2)
    data_deal_line_3 = np.array(data_deal_line_3)
    data_deal_line_4 = np.array(data_deal_line_4)

    data_deal_line = np.concatenate(
        (data_deal_line_1, data_deal_line_2, data_deal_line_3, data_deal_line_4), axis=1)

    data_deal_line = pd.DataFrame(data_deal_line)
    columns_line = data_deal_line.columns

    l_line = len(columns_line)
    for i in range(data_deal_line.shape[1]):
        lie1 = data_deal_line[columns_line[i]]
        ws1.cell(1, i + 1, str(columns_line[i] + 1))
        for ii in range(len(lie1)):
            ws1.cell(ii + 2, l_line + 1, cos[ii])
            ws1.cell(ii + 2, i + 1, lie1[ii])
    ws1.cell(1, l_line + 1, l_line + 1)
    wb1.save(end_file_name2)
    to_pkl(end_file_name2)

    # peak_1_data
    data_deal_peak_1 = pd.DataFrame(data)

    data_deal_peak_1 = data_deal_peak_1.iloc[:, line_1_shape:line_1_shape + peak_1_shape]

    columns_peak1 = data_deal_peak_1.columns
    l_peak1 = len(columns_peak1)
    for i in range(data_deal_peak_1.shape[1]):
        lie1 = data_deal_peak_1[columns_peak1[i]]
        ws2_1.cell(1, i + 1, i + 1)
        for ii in range(len(lie1)):
            ws2_1.cell(ii + 2, l_peak1 + 1, cos[ii])
            ws2_1.cell(ii + 2, i + 1, lie1[ii])
    ws2_1.cell(1, l_peak1 + 1, l_peak1 + 1)
    wb2_1.save(end_file_name3_1)
    to_pkl(end_file_name3_1)

    # peak_2_data
    data_deal_peak_2 = pd.DataFrame(data)
    data_deal_peak_2 = data_deal_peak_2.iloc[:,
                       line_1_shape + peak_1_shape + line_2_shape:line_1_shape + peak_1_shape + line_2_shape + peak_2_shape]
    columns_peak2 = data_deal_peak_2.columns
    l_peak2 = len(columns_peak2)
    for i in range(data_deal_peak_2.shape[1]):
        lie1 = data_deal_peak_2[columns_peak2[i]]
        ws2_2.cell(1, i + 1, i + 1)
        for ii in range(len(lie1)):
            ws2_2.cell(ii + 2, l_peak2 + 1, cos[ii])
            ws2_2.cell(ii + 2, i + 1, lie1[ii])
    ws2_2.cell(1, l_peak2 + 1, l_peak2 + 1)
    wb2_2.save(end_file_name3_2)
    to_pkl(end_file_name3_2)

    # peak_3_data
    data_deal_peak_3 = pd.DataFrame(data)
    data_deal_peak_3 = data_deal_peak_3.iloc[:,
                       line_1_shape + peak_1_shape + line_2_shape + peak_2_shape + line_3_shape:line_1_shape + peak_1_shape + line_2_shape + peak_2_shape + line_3_shape + peak_3_shape]
    columns_peak3 = data_deal_peak_3.columns
    l_peak3 = len(columns_peak3)
    for i in range(data_deal_peak_3.shape[1]):
        lie1 = data_deal_peak_3[columns_peak3[i]]
        ws2_3.cell(1, i + 1, i + 1)
        for ii in range(len(lie1)):
            ws2_3.cell(ii + 2, l_peak3 + 1, cos[ii])
            ws2_3.cell(ii + 2, i + 1, lie1[ii])
    ws2_3.cell(1, l_peak3 + 1, l_peak3 + 1)
    wb2_3.save(end_file_name3_3)
    to_pkl(end_file_name3_3)

    # peak_data
    p1 = np.array(data_deal_peak_1)
    p2 = np.array(data_deal_peak_2)
    p3 = np.array(data_deal_peak_3)
    deal_peak_all = np.concatenate((p1, p2, p3), axis=1)
    deal_peak_all_pd = pd.DataFrame(deal_peak_all)
    columns_peak = deal_peak_all_pd.columns
    l_peak = len(columns_peak)
    for i in range(deal_peak_all_pd.shape[1]):
        lie1 = deal_peak_all_pd[columns_peak[i]]
        ws2.cell(1, i + 1, i + 1)
        for ii in range(len(lie1)):
            ws2.cell(ii + 2, l_peak + 1, cos[ii])
            ws2.cell(ii + 2, i + 1, lie1[ii])
    ws2.cell(1, l_peak + 1, l_peak + 1)
    wb2.save(end_file_name3)
    to_pkl(end_file_name3)
    print("Dataset creation completed!")


def dwt(x, wavelet='db1'):
    data_x = []
    for i in x:
        cA, cD = pywt.dwt(i, wavelet, mode='symmetric')
        ya = pywt.idwt(cA, None, wavelet, mode='symmetric')
        # yd = pywt.idwt(None, cD, wavelet, mode='symmetric')
        data_x.append(ya)
    return np.array(data_x)


def Dwt_data(path, end_file_name1):
    wb_all = Workbook()
    wb_all.create_sheet(index=0, title="all")
    ws_all = wb_all.active
    # data = pd.read_pickle(path)
    data = pd.read_excel(path)  # test 50
    data = pd.DataFrame(data)
    c_value = data.iloc[:, -1]
    data = data.iloc[:, 0:-1]
    arr_data = np.array(data)
    dwt_data = dwt(arr_data)
    deal_data_line_all_dwt = pd.DataFrame(dwt_data)
    columns_ = deal_data_line_all_dwt.columns
    # all_data
    l = len(columns_)
    i1 = 0
    for i in range(l):
        i1 += 1
        lie = deal_data_line_all_dwt[columns_[i]]
        for j in range(len(lie)):
            ws_all.cell(1, i + 1, i1)
            ws_all.cell(j + 2, i + 1, lie[j])
    for col_i in range(len(c_value)):
        col = c_value[col_i]
        ws_all.cell(col_i + 2, l + 1, col)
    ws_all.cell(1, l + 1, l + 1)
    name = end_file_name1
    wb_all.save(name)
    path = to_pkl(end_file_name1)
    print("Data DWT completed!")
    return path


def run():
    oldpath = 'Data_human'  # Original Data Location
    newpath = oldpath + "_deal_dwt"  # Generate Process Data Location
    xlsx_data = oldpath + "_xlsx_dwt"
    end_file_name0 = xlsx_data + '/line_peak.xlsx'  # Final generation of dataset catalog
    end_file_name1 = xlsx_data + '/line_peak_dwt.xlsx'  # Final generation of dataset catalog
    end_file_name2 = xlsx_data + '/lines_dwt.xlsx'  # Final generation of dataset catalog
    end_file_name3 = xlsx_data + '/peaks_dwt.xlsx'  # Final generation of dataset catalog
    end_file_name3_1 = xlsx_data + '/peak_1_dwt.xlsx'  # Final generation of dataset catalog
    end_file_name3_2 = xlsx_data + '/peak_2_dwt.xlsx'  # Final generation of dataset catalog
    end_file_name3_3 = xlsx_data + '/peak_3_dwt.xlsx'  # Final generation of dataset catalog

    txt_num = 1 # Average the last number of txt files
    nn = 9
    del_files(newpath, xlsx_data)
    read_data(oldpath, newpath, xlsx_data, txt_num)
    path1 = remove_bd(newpath)
    path2 = writeinexcel(path1, nn)

    path3, line_1_shape, line_2_shape, line_3_shape, line_4_shape, peak_1_shape, peak_2_shape, peak_3_shape = dif(path2)
    # path4 = Extended_data(path3, end_file_name0)
    path5 = Dwt_data(path3, end_file_name1)
    Distribute_data(path5, end_file_name2, end_file_name3, end_file_name3_1, end_file_name3_2,
                    end_file_name3_3, 74, 151, 180, 51, 37, 38, 39)
    return xlsx_data
